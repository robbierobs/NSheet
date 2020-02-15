# LMIS Enola ITs hyperlink
# LMIS Shop by reason https://www2.nscorp.com/mech0000/unitshopreason.lmis
# LMIS Shop By Reason units
# https://www2.nscorp.com/mech0000/unitshopreasonitdetail.lmis?Shop=ENO&Reason=
# <a onclick="javascript:openNewWindow('Count','ENO','IT');"
# style="cursor:pointer;">6</a>

# Lets create a database (SQLite) to keep track of employees (NSure),
# locomotives, fuel, cs, dp, lsl, direction, etc. 

import sqlite3
from datetime import datetime
from datetime import date
from datetime import timedelta
from openpyxl import load_workbook
from openpyxl import Workbook
from lxml import html
from bs4 import BeautifulSoup
from os import path
import getpass
import csv
import os
import sys
import shutil
import requests

current_date_time = datetime.today()
recorded_time = current_date_time.strftime("%m-%d-%Y %H:%M:%S")
today = date.today()
curDate = today.strftime("%m-%d-%Y")
fromList = ' '
#wb = load_workbook(filename = 'Enola_Powersheet_'+curDate+'.xlsx')
wb = load_workbook(filename = 'Enola Powersheet 12-14-2019 1430.xlsx')
#current_sheet = wb[curDate]
current_sheet = wb['12-14-2019']
locomotive_dictionary = {}

locomotive_notes_date = 'locomotive_notes/'+curDate

#   This populates the locomotive_dictionary variable with data that may be
#   left behind after a crash or leaving the program
if os.path.exists(locomotive_notes_date) is True:
    #keep_old = input('Would you like to import the backup data? ')
    keep_old = 'y'
    if keep_old == 'y' or keep_old == 'Y':
        found_files = [name for name in os.listdir(os.path.join(locomotive_notes_date)) if
            os.path.isfile(os.path.join(locomotive_notes_date, name))]
        for found in found_files:
            if found not in locomotive_dictionary.keys():
                with open(os.path.join(locomotive_notes_date, found), 'r') as f:
                    found_loco_notes = [line.strip() for line in f]
                    for lines in found_loco_notes:
                        if found not in locomotive_dictionary.keys():
                            locomotive_dictionary[found] = []
                            locomotive_dictionary[found].append(lines)
                        else:
                            locomotive_dictionary[found].append(lines)

locomotive_note_source='./locomotive_notes/'+curDate
locomotive_note_backup='./locomotive_backup'

if os.path.exists(locomotive_note_source):
    #remove_backup_input = input('Would you like to delete the backup information? ')
    remove_backup_input = 'n'
    if remove_backup_input == 'y' or remove_backup_input == 'Y':
        shutil.rmtree(locomotive_note_backup)
        shutil.move(locomotive_note_source, locomotive_note_backup)
#    os.rmdir(locomotive_notes_date)
if os.path.exists(locomotive_notes_date) is False:
    os.mkdir('locomotive_notes/'+curDate)

conn = sqlite3.connect('enola_database.db')
c = conn.cursor()

#print('Currently working with the',curDate,'worksheet...')
print('Using a static sheet currently')
print('At the moment this is only being built for C-trick')
print('Add a turnover function to automatically fill that out\n\n')
print('FIXME:  3 monthers are not added correctly to work packet.')

print("************-| LD_50 Locomotive Powersheet Mutilator |-**************")
def menu():

    choice = input("""
                A: Inbound and Outbound Report
                B: Outbound Trains
                C: Change Powersheet
                D: Save Powersheet
                E: Rob Peter
                F: Pay Paul
                G: Search for open engines
                H: Identify shoppers
                I: Make Work Packets
                J: Scrape Unit Information (needs updated)
                K: Locomotive Notes
                L: Rundown

                Q: Quit/Log Out

                Please enter your choice: """)

    if choice == "A" or choice == "a":
        dispatchReport()
    elif choice == "B" or choice == "b":
        fromBuilt()
    elif choice == "C" or choice == "c":
        change_powersheet()
    elif choice == "D" or choice == "d":
        savePowersheet()
        menu()
    elif choice == "E" or choice == "e":
        robPeter()
    elif choice == "F" or choice == "f":
        payPaul()
    elif choice == 'G' or choice == 'g':
        openEngines()
        menu()
    elif choice == 'H' or choice == 'h':
        shoppers()
    elif choice == 'I' or choice == 'i':
        create_packets()
    elif choice == 'J' or choice == 'j':
        scrape()
    elif choice == 'K' or choice == 'k':
        notes()
    elif choice == 'L' or choice == 'l':
        rundown()
        menu()
    elif choice == "Q" or choice == "q":
        sys.exit
    else:
        print("You must only select either A,B,C,D,E,F,G or Q.")
        print("Please try again\n")
        menu()
def change_powersheet():
    change_sheet = input('''
                A: Change Inbound Powersheet
                B: Change Outbound Powersheet
                C: Change Inbound Pit Sheet
                D: Change Turnover Sheet

                Please enter your choice:  ''')
    if change_sheet == 'a' or change_sheet == 'A':
        appendInbound()
    elif change_sheet == 'b' or change_sheet == 'B':
        appendBuild()
    else:
        print('FIXME:Other functions still need built.')
        menu()
def create_packets():
    USERNAME = input('\nLMIS Username: ')
    PASSWORD = getpass.getpass('LMIS Password: ')
    #USERNAME = 'iywaa'
    #PASSWORD = ''
    print('LD_50 Scrape')
    print('Author: Sean Robinson, SGL, Enola Diesel')
    print('Welcome to the LMIS Scraper...\n')

    locomotive_numbers = input('Enter locomotive numbers separated by a space: ')
    locomotive_list = locomotive_numbers.split(" ")
    date = today.strftime("%m-%d-%Y")
    unitInfo = []
    scheduled = []
    scheduled_date = []
    scheduled_tasks = []
    scheduled_task_dates = []
    # login page for LMIS
    LOGIN_URL = "https://www2.nscorp.com/mech0000/login.lmis"

    payload = {
        "username": USERNAME, 
        "pass1": PASSWORD, 
    }

    # keeps us logged into the session
    session_requests = requests.session()
    result = session_requests.get(LOGIN_URL)

    # Login
    result = session_requests.post(LOGIN_URL, data = payload, headers = dict(referer = LOGIN_URL))
     
    # Loop over the input list and scrape the work orders
    for x in locomotive_list:
        # unit_in_shop_by_reason = ('https://www2.nscorp.com/mech0000/unitshopreason.lmis')
        # shop_it_units = ('https://www2.nscorp.com/mech0000/unitshopreasonitdetail.lmis?Shop=ENO&Reason=')
        LMIS_URL = ("https://www2.nscorp.com/mech0000/OutstandingWorkOrders.lmis?pageprocess=VT&locoinit=NS&loconbr=000000"+x+"&notfromshp=N&readonly=N&shop=%20%20%20&attachonly=N&updateact=N&searchbox=Y&reqFromModule=")
        scheduled_dates_url ="https://www2.nscorp.com/mech0000/SmDueDates.lmis?action=S&callingScreen=OUTWRKOR&unitinit=NS&unitnumber=000000"+x+"&inclsmi=N"
        result = session_requests.get(LMIS_URL, headers = dict(referer = LMIS_URL))
        scheduled_result = session_requests.get(scheduled_dates_url, headers = dict(referer = scheduled_dates_url))

        # unit_result = session_requests.get(unit_information_report, headers = dict(referer = unit_information_report))
        soup = BeautifulSoup(result.content, 'lxml')
        scheduled_soup = BeautifulSoup(scheduled_result.content, 'lxml')

        # unit_soup = BeautifulSoup(unit_result.content, 'lxml')

        # add PTC health when able
        # add DP also
        # scheduled due dates 
        # items are hidden on the due dates page
        # they are not static and will have to be searched through

        #this print the category and due date, make this into a list or table?
        for task in range(0, 19):
            due_task = scheduled_soup.find("input", {"name":"hTaskType"+str(task)})['value']
            due_date_str = scheduled_soup.find("input", {"name":"hNextDueDate"+str(task)})['value']
            due_date = datetime.strptime(due_date_str, '%m-%d-%Y')
            mi_date = (datetime.today() + timedelta(6))
            if due_date < mi_date:
                scheduled_tasks.append(due_task)
                scheduled_task_dates.append(due_date_str)
            #compare dates and add due dates to a list
            #check list for comparison to see if something exists (lube, labs, af, etc)
            #if it exists, make a variable so it can be added to the work packet
        scheduled.append(list(scheduled_tasks))
        scheduled_date.append(list(scheduled_task_dates))
        scheduled_tasks.clear()
        scheduled_task_dates.clear()
        # TS,LS,LB,1Y,2Y,N6,M5,M6,AF,EV,MR,CS,HB,AN,AB,RS, N2,N4,N5,M7,M2,M3,M5,RS(tape)

        print('FIXME: add ptc health to end of PTC line...add DP to report\n')
        print('----',x,'----')
        if (soup.find("input", {"name":"hModel"})) is not None:
            model = soup.find("input", {"name":"hModel"})['value']
            print('Model: ' + model)
        else:
            model = '-'
        if (soup.find("input", {"name":"hPtc"})) is not None:
            ptc = soup.find("input", {"name":"hPtc"})['value']
            print('PTC: ' + ptc)
        else:
            ptc = '-'
        if (soup.find("input", {"name":"hEM"})) is not None:
            em = soup.find("input", {"name":"hEM"})['value']
            print('EM: ' + em)
        else:
            em = '-'
        if (soup.find("input", {"name":"hCs"})) is not None:
            cabs = soup.find("input", {"name":"hCs"})['value']
            print('CS: ' + cabs)
        else:
            cabs = '-'
        if (soup.find("input", {"name":"hLSL"})) is not None:
            lsl = soup.find("input", {"name":"hLSL"})['value']
            print('LSL: ' + lsl)
        else:
            lsl = '-'
        if (soup.find("input", {"name":"hRelIu"})) is not None:
            relInd = soup.find("input", {"name":"hRelIu"})['value']
            print('Reliability: ' + relInd)
        else:
            relInd = '-'
        if (soup.find("input", {"name":"hEquivAxl"})) is not None:
            group = soup.find("input", {"name":"hEquivAxl"})['value']
            print('Power Group: ' + group)
        else:
            group = '-'
        if (soup.find("input", {"name":"hPropDue"})) is not None:
            fra = soup.find("input", {"name":"hPropDue"})['value']
            print('FRA Due: ' + fra)
        else:
            fra = '-'
        if (soup.find("input", {"name":"hEpaDead"})) is not None:
            epa = soup.find("input", {"name":"hEpaDead"})['value']
            print('EPA Due: ' + epa)
        else:
            epa = '-'
        if (soup.find("input", {"name":"hLubeDue"})) is not None:
            lube = soup.find("input", {"name":"hLubeDue"})['value']
            print('Lube Due: ' + lube)
        else:
            lube = '-'
        if (soup.find("input", {"name":"hCabS"})) is not None:
            csDue = soup.find("input", {"name":"hCabS"})['value']
            print('Cab Signals Due: ' + csDue)
        else:
            csDue = '-'
        if (soup.find("input", {"name":"hFc"})) is not None:
            fuelCap = soup.find("input", {"name":"hFc"})['value']
            fuel_capacity = fuelCap.lstrip("0")
            print('Fuel Capacity: ' + fuel_capacity)
        else:
            fuelCap = '-'
        if (soup.find("input", {"name":"hNextFraAirFlowMeter"})) is not None:
            airFlow = soup.find("input", {"name":"hNextFraAirFlowMeter"})['value'] 
        else:
            airFlow = '-'
        if (soup.find("input", {"name":"hHomeShpCd"})) is not None:
            home = soup.find("input", {"name":"hHomeShpCd"})['value'] 
        else:
            home = '-'
        if (soup.find("input", {"name":"hAltrShpCd"})) is not None:
            alt = soup.find("input", {"name":"hAltrShpCd"})['value'] 
        else:
            alt = '-'
 

        # if statement that has assigned engine in the mix.
        # We can use this to aid in automatically populating the inbound
        # turnover sheet.

        locomotive_Info = x,date,fra,epa,'Y',csDue
        unit_num = int(x)
        
        # Eventually this command will need to be replaced to only insert the
        # correct values, this was built as a test for now and will be used
        # accordingly until SQL data can be pulled from another database or
        # until I leave them blank to be inputted to other 

        c.execute("INSERT INTO enoladb (unit_number, model, cab_signals, lsl,\
                  fra_date, epa_date, lube_due, cs_due, afm_due,\
                  fuel_capacity, updated, home_shop, alt_shop) VALUES\
                  (?,?,?,?,?,?,?,?,?,?,?,?,?)\
                  ON CONFLICT(unit_number) DO UPDATE SET \
                  fra_date=excluded.fra_date,\
                  epa_date=excluded.epa_date,\
                  lube_due=excluded.lube_due,\
                  cs_due=excluded.cs_due,\
                  afm_due=excluded.afm_due,\
                  updated=excluded.updated,\
                  home_shop=excluded.home_shop,\
                  alt_shop=excluded.alt_shop;",
                  (x,model,cabs,lsl,fra,epa,lube,csDue,airFlow,fuel_capacity,\
                   recorded_time,home,alt))
 
        conn.commit()
        unitInfo.append(locomotive_Info)

    correctInfo=input("\nIs the information correct? (y/n) ")
    if correctInfo == 'y' or correctInfo == 'Y':
        mi_starting_cell = 25
        ur_starting_cell = 23
        miCover = load_workbook(filename="MIPacketCover.xlsx")
        urCover = load_workbook(filename="URPacketCover.xlsx")
        j = 0
        for info in unitInfo:
            print('\n|--------| '+info[0]+' |----------|\n')
            table_format(scheduled[int(j)], scheduled_date[int(j)],'Task','Due Date')
            maint = input('\nIs '+info[0]+' a maintenance unit? (y/n) ')
            if maint == "y" or maint == "Y":
                print('Saving cover for Unit #: '+info[0]+'.')
                packet = miCover.copy_worksheet(miCover["MI Cover Sheet"])
                packet.title=info[0]
                maintenance_dates(scheduled[int(j)], scheduled_date[int(j)],
                                 packet)
                worksheet_tasks(packet, mi_starting_cell, info[0])
                packet.cell(row=2, column=1).value = info[0]
                #packet.cell(row=1, column=6).value = info[1]
                packet.cell(row=5, column=3).value = info[2]
                packet.cell(row=4, column=6).value = 'Y'
                j += 1
            elif maint == 'n' or maint == 'N':
                print('Saving cover for Unit #: '+info[0]+'.')
                packet = urCover.copy_worksheet(urCover["UR Cover Sheet"])
                packet.title=info[0]
                packet.cell(row=2, column=1).value = info[0]
                packet.cell(row=1, column=6).value = info[1]
                packet.cell(row=5, column=3).value = info[2]
                maintenance_dates(scheduled[int(j)], scheduled_date[int(j)],
                                  packet)
                worksheet_tasks(packet, ur_starting_cell, info[0])
                packet.cell(row=4, column=6).value = 'Y'
                j += 1
        urCover.save('UR_CoverSheets.xlsx')
        miCover.save('MI_CoverSheets.xlsx')
    menu()

def worksheet_tasks(packet, cell, loco_number):
    work_list = []
    work_header = []

    #searching dictionary to see if notes exist for the engine
    #if notes exist, prmopt to add the note to the packet

    if loco_number in locomotive_dictionary.keys():
        print('Found it!!!')
        for values in locomotive_dictionary[loco_number]:
            print(values)
            max_string_value = '%.20s' % values
            add_to_work_packet = input('Would you like to add "'
                                       +max_string_value+'" to the work packet?' )
            if add_to_work_packet == 'y' or add_to_work_packet == 'Y':
                dictionary_header = input('Enter a header for this task: ')
                print('Adding to the packet.')
                work_list.append(values)
                work_header.append(dictionary_header)

    while True:
        header = input('Input worksheet header: ')
        if header == 'Q' or header =='q' or header == '':
            break
        work_header.append(header)
        print(work_header)
        work_task = input('Input worksheet task: ')
        if work_task == 'Q' or work_task == 'q' or work_task == '':
            break
        work_list.append(work_task)
        print(work_list)

    for head in work_header:
        work_cell_iterator = int(cell)
        while packet.cell(row=work_cell_iterator, column=2).value is not None:
            if work_cell_iterator == 32:
                work_cell_iterator = int(work_cell_iterator) + 6
            elif work_cell_iterator == 33:
                work_cell_iterator = int(work_cell_iterator) + 5
            else:
                work_cell_iterator = int(work_cell_iterator) + 3
        print('Adding header: '+head)
        packet.cell(row=work_cell_iterator, column=2).value = head

    for task in work_list:
        work_cell_iterator = int(cell) + 1
        while packet.cell(row=work_cell_iterator, column=2).value is not None:
            if work_cell_iterator == 33:
                work_cell_iterator = int(work_cell_iterator) + 6
            elif work_cell_iterator == 34:
                work_cell_iterator = int(work_cell_iterator) + 5
            else:
                work_cell_iterator = int(work_cell_iterator) + 3
        print('Adding task: '+task)
        packet.cell(row=work_cell_iterator, column=2).value = task

def maintenance_dates(tasks, due_dates, packet):
    mi_due_cell = packet.cell(row=6, column=3).value
    epa_due_cell = packet.cell(row=7, column=3).value
    epa_due = ''
    mi_due = ''
    if 'LS' in tasks:
        task_index = tasks.index('LS')
        packet.cell(row=3, column=6).value = 'Y'
    if 'AF' in tasks:
        task_index = tasks.index('AF')
        packet.cell(row=7, column=6).value = 'Y'
    if 'CS' in tasks:
        packet.cell(row=6, column=6).value = 'Y'
    if 'LB' in tasks:
        packet.cell(row=5, column=6).value = 'Y'
    if '1Y' in tasks and '2Y' in tasks:
        packet.cell(row=7, column=3).value = epa_due+'1Y, 2Y'
    elif '1Y' in tasks:
        packet.cell(row=7, column=3).value = epa_due+'1Y'
    elif '2Y' in tasks:
        packet.cell(row=7, column=3).value = epa_due+'2Y'
    if 'M5' in tasks:
        packet.cell(row=7, column=3).value = epa_due+'M5'
    if 'M6' in tasks:
        packet.cell(row=7, column=3).value = epa_due+'M6'
    if 'M7' in tasks:
        packet.cell(row=7, column=3).value = epa_due+'M7'
    if 'N6' in tasks:
        packet.cell(row=7, column=3).value = epa_due+',N6'
    if 'MR' in tasks:
        packet.cell(row=6, column=3).value = '6mo'
    if 'AN' in tasks:
        packet.cell(row=6, column=3).value = '12mo'
    if 'AB' in tasks:
        packet.cell(row=6, column=3).value = packet.cell(row=6,
                                                         column=3).value+', Air'
def writeMultColumns(row, column1, column2, robbed, paid):
    r = row
    col1 = column1
    col2 = column2
    t = robbed
    p = paid
    #u = units

    print('|---------- Writing Data ----------|')

    #check for empty cells
    emptyCol1 = current_sheet.cell(row=r, column=col1).value
    emptyCol2 = current_sheet.cell(row=r, column=col2).value 

    if emptyCol1 and emptyCol2 is None:
        emptyCol1 = t
        emptyCol2 = p

def notes():
    # This function will allow you to take notes of whatever power you would
    # like, this allows you to search power at the beginning of the shift, look
    # them up throughout the shift and finally use the notes in the creation of
    # work packets if need be.

    #print('\nFIXME: Add list of current notes on startup of this feature.')
    #print('\nFIXME: Add commands list to remind, D to delete notes,\n'+
    #      'V to view list of current dictionary')
    print('Current notes available for the following locomotives: ')
    # To keep the locomotives ordered and immutable, we are going to use a
    # dictionary. This allows key-value pairs that would be easier to search
    # through and less likely to be corrupted compared to a list
    #for key in locomotive_dictionary.keys():
    #    print(key)
    locomotive = input('\nEnter locomotive number: ')
    if locomotive == 'v' or locomotive ==  'V':
        for key in locomotive_dictionary.keys():
            print(key)
    elif locomotive == '' or locomotive == 'q' or locomotive == 'Q':
        print('Exiting.')
    elif locomotive == 'del all':
        locomotive_dictionary.clear()
    elif locomotive in locomotive_dictionary:
        print("Locomtive notes already exist. Adding to current notes...")
        print("(v) to view notes, (d) delete notes, (q) to exit")
        while True:
            note = input('Enter note: ')
            if note == '' or note == 'Q' or note == 'q':
                print(locomotive_dictionary)
                break
            elif note == 'D' or note == 'del':
                print('Deleting notes for locomotive', locomotive)
                del locomotive_dictionary[locomotive]
                break
            elif note == 'V' or note =='v':
                print(locomotive_dictionary[locomotive])
            else:
                locomotive_dictionary[locomotive].append(note)
    else:
        locomotive_dictionary[locomotive] = []
        while True:
            note = input('Enter note: ')
            if note == '' or note == 'Q' or note == 'q':
                print(locomotive_dictionary)
                break
            elif note == 'D':
                print('Deleting notes for locomotive', locomotive)
                del locomotive_dictionary[locomotive]
                break
            else:
                locomotive_dictionary[locomotive].append(note)
                #print(locomotive_dictionary)
                if os.path.exists(locomotive_notes_date) is False:
                    os.mkdir('locomotive_notes/'+curDate)
                with open(os.path.join('locomotive_notes/'+curDate,locomotive), 'w', newline='') as loco_text:
                    for notes in locomotive_dictionary[locomotive]:
                        loco_text.write(notes+'\r\n')
    menu()

def readSingleColumn(rowStart, rowEnd, col):
    row_start = rowStart
    row_end= rowEnd
    column = col
    this_list= []
    for row in range(row_start, row_end):
        cell_value = current_sheet.cell(row=row, column=column).value
        if cell_value is not None:
            this_list.append(cell_value)
        elif cell_value is None:
            this_list.append('-')
    return this_list

def readMultColumns(rowStart, rowEnd, colStart, colEnd):
    rs = rowStart
    re = rowEnd
    cs = colStart
    ce = colEnd
    for i in range(rs, re):
        x = current_sheet.cell(row=i, column=cs).value
        z = current_sheet.cell(row=i, column=ce).value
        if x is None and z is not None:
            print(x,z)
        elif x is int and x is not None:
            print(x,'-->',z)
        elif x is not None:
            print(str(x)[:3],'-->',z)
            #print(x,'-->',z)

def readMultColumnsTable(rowStart, rowEnd, colOne, colTwo):
    row_start = rowStart
    row_end = rowEnd
    column_one = colOne
    column_two = colTwo
    list_one = []
    list_two = []

    for cell in range(row_start, row_end):
        column1 = current_sheet.cell(row=cell, column=column_one).value
        column2 = current_sheet.cell(row=cell, column=column_two).value
        if column1 is None and column2 is not None:
            list_one.append('-')
            list_two.append(column2)
        elif column1 is int and column2 is not None:
            list_one.append(column1)
            list_two.append(column2)
        elif column1 is not None and column2 is None:
            list_one.append(str(column1)[:3]+' ---------->')
            list_two.append('-')
        elif column1 is not None:
            list_one.append(str(column1)[:3]+' ---------->')
            list_two.append(column2)
    return list_one, list_two

def table_format(list_one, list_two, label_one, label_two):
    fmt = '{:<8}{:<20}{}'
    print(fmt.format('', label_one, label_two))
    for stuff, (row,column) in enumerate(zip(list_one, list_two)):
        print(fmt.format((stuff + 1), row, column))
        
def table_format_three(list_one, list_two, list_three,
                       label_one, label_two, label_three):
    fmt = '{:<8}{:<20}{:<50}{}'
    print(fmt.format('', label_one, label_two, label_three))
    for stuff, (list_a, list_b, list_c) in enumerate(zip(list_one, list_two, list_three)):
        print(fmt.format((stuff + 1), list_a, list_b, list_c))

def openEngines():
    print('\nSearching for Open Locomotives on Current Sheet.....')
    inboundUnits=[]
    for rows in range (4, 27):
        row = current_sheet.cell(row=rows, column = 3).value
        if row is not None:
            rowSlicer = row.split()
            slash = '/'
            while slash in rowSlicer: rowSlicer.remove(slash)
            newList = [s[:4] for s in rowSlicer if s[:3].isdigit()]
            inboundUnits.extend(newList)
    usedUnits=[]
    for rows in range(4, 27):
        row = current_sheet.cell(row=rows, column=14).value
        if row is not None:
            rowSlicer = row.split()
            slash = '/'
            while slash in rowSlicer: rowSlicer.remove(slash)
            newList = [s[:4] for s in rowSlicer if s[:3].isdigit()]
            usedUnits.extend(newList)

    openUnits = [s for s in inboundUnits if s not in usedUnits]
    print('Search complete...\n\nOpen Locomtovies:',openUnits,'\n')

def robPeter():
    print('\n|--------------- Robbing Peter ---------------|\n')
    stolenTrain = input('What train are we stealing from?  ')
    stolenUnits = input('What units are being stolen?  ')
    stolenList = stolenUnits.split()
    print(stolenList)
    payTrain = input('What train are we paying '+stolenTrain+' to? (ex. 15T - OUTBOUND)  ')

    #check inbound side for an empty slot to fill in the robbed train
    #then insert the information 
    for train in range(15, 28):
        checkTrain = current_sheet.cell(row=train, column=1).value
        checkUnits = current_sheet.cell(row=train, column=3).value
        if checkTrain is None and checkUnits is None:
            if stolenTrain.find('.') >= 0:
                datedTrain = stolenTrain.split('.')
                current_sheet.cell(row=train,column=1).value=datedTrain[0]
                current_sheet.cell(row=train, column=2).value=datedTrain[1]
                current_sheet.cell(row=train, column=3).value = stolenUnits
            else:
                current_sheet.cell(row=train, column=1).value = stolenTrain
                current_sheet.cell(row=train, column=3).value = stolenUnits
            current_sheet.cell(row=train, column=8).value = payTrain
            print(checkTrain,':',checkUnits)
            break

    #print('How many units will be needed to replace robbed power?')
    needed=input('How many units will be needed to replace robbed power? (NEED X)  ')
    
    #check for empty rows in the outbound extras 
    for train in range(15, 28):
        checkTrain = current_sheet.cell(row=train, column=12).value
        checkUnits = current_sheet.cell(row=train, column=14).value
        if checkTrain is None and checkUnits is None:
            current_sheet.cell(row=train, column=12).value = stolenTrain
            current_sheet.cell(row=train, column=14).value = needed
            print(checkTrain,':',checkUnits)
            break
    dispatchReport()

def payPaul():
    print('\n|--------------- Paying Paul ---------------|\n')
    
    payTrain = input('What train are we paying back?  ')
    payUnits = input('What units are being given?  ')

    payList = payUnits.split()
    print(payUnits)
    
    fromTrain = input('What train are we paying '+payTrain+' from? ')

    for train in range(16, 28):
        checkTrain = current_sheet.cell(row=train, column=12).value
        checkUnits = current_sheet.cell(row=train, column=14).value
        if checkTrain == payTrain:
            current_sheet.cell(row=train, column=14).value = payUnits
            current_sheet.cell(row=train, column=19).value = fromTrain
            print(checkTrain,':',checkUnits)
            break
    #print('Repaid check')

    dispatchReport()

def shoppers():
    shoppers = input('\nEnter shopped units coming in: ')
    print(shoppers)
    print('FIXME: Use this function to color shoppers red and add them to the assignemnts sheets')
    menu()

def searchPower(newPower, fromRow):
    power = [x.strip() for x in newPower.split('/')]
    row=fromRow
    units=[]
    for i in power:
        x = i[:4]
        for j in range (4, 27):
            # assigning y to train symbols
            y = current_sheet.cell(row=j, column=1).value
            if y is not None:
                a = str(y)[:3]
                z = current_sheet.cell(row=j, column=3).value
                if z is not None:
                    zlist = z.strip()
                    search = zlist.find(x)
                    if search >= 0:
                        units.append(a)
    remove_adjacent(units)
    fromList = (' / '.join(units))
    return fromList
    #print('From: ',fromList)
    #current_sheet.cell(row=row, column=21).value = fromList

def remove_adjacent(seq): # works on any sequence, not just on numbers
    i = 1
    n = len(seq)
    while i < n: # avoid calling len(seq) each time around
        if seq[i] == seq[i-1]:
            del seq[i]
            # value returned by seq.pop(i) is ignored; slower than del seq[i]
            n -= 1
        else:
            i += 1
def appendInbound():
    print('Add function to change inbound power, search for maxes and update oubound sheets like the oteher function.')
    print('\n|*************** Changing Power ***************|\n')

    print('\n----------| Inbound Trains |------------\n')
    inbound_train,inbound_power = readMultColumnsTable(4, 16, 1, 3)
    table_format(inbound_train, inbound_power, 'Train Symbol', 'Power')
    selectPower = input('\n\nWhat train would you like to change?  ')
    # row will be the current row, we have to add 3 to get down to the correct cell
 
    row = int(selectPower) + int(3)
    old_build = current_sheet.cell(row=int(row), column=3).value
    print('\n|---------- Appending build for',current_sheet.cell(row=row,column=1).value,'----------|\n')
    print('Current build: ', old_build)
    newBuild = input('New build:  ')
    confirmBuild = input('You would like to change the build to: ' + newBuild + ' ? (y/n)  ')
    if confirmBuild == 'y' or confirmBuild =="Y":
        print('\nChanging power from ', old_build, ' to :', newBuild,' ')
        current_sheet.cell(row=row, column=3).value = newBuild
        print('Power changed to', current_sheet.cell(row=row, column=3).value, ' \n')
        #Editing the "From" category on Outbound Trains
        if newBuild == '' or newBuild == ' ' or newBuild == 'None':
            newBuildValue = ' '
            print('Removing build from ',selectPower+'.')
            current_sheet.cell(row=row, column=21).value = ''
        else:
            newBuildValue = current_sheet.cell(row=row, column=3).value
            print('Consist:',newBuildValue)
            #Searches the inbound columns to find where power came from
            foundPower = searchPower(newBuildValue, row)
            current_sheet.cell(row=row, column=21).value = foundPower
            print(current_sheet.cell(row=row, column=1).value,' changed to: ',newBuild, ' FROM:', current_sheet.cell(row=row, column=21).value, '\n')
    else:
        print('No changes made.\n')
    menu()


def appendBuild():
    print('\n|*************** Changing Power ***************|\n')

    print('\n----------| Inbound Trains |------------\n')
    inbound_train,inbound_power = readMultColumnsTable(4, 16, 1, 3)
    table_format(inbound_train, inbound_power, 'Train Symbol', 'Power')

    print('\n----------| Outbound Trains |------------\n')
    outbound_train, outbound_power = readMultColumnsTable(4, 16, 12, 14)
    table_format(outbound_train, outbound_power, 'Train Symbol', 'Power')

    selectPower = input('\n\nWhat train would you like to change?  ')
    # row will be the current row, we have to add 3 to get down to the correct cell
 
    row = int(selectPower) + int(3)
    old_build = current_sheet.cell(row=int(row), column=14).value
    openEngines()
    print('\n|---------- Appending build for',current_sheet.cell(row=row, column=12).value,'----------|\n')
    print('Current build: ', old_build)
    newBuild = input('New build:  ')
    confirmBuild = input('You would like to change the build to: ' + newBuild + ' ? (y/n)  ')

    if confirmBuild == 'y' or confirmBuild =="Y":
        print('\nChanging power from ', old_build, ' to :', newBuild,' .')
        current_sheet.cell(row=row, column=14).value = newBuild
        print('Power changed to', current_sheet.cell(row=row, column=14).value, ' .\n')
        #Editing the "From" category on Outbound Trains
        if newBuild == '' or newBuild == ' ' or newBuild == 'None':
            newBuildValue = ' '
            print('Removing build from ',current_sheet.cell(row=row, column=12).value+'.')
            current_sheet.cell(row=row, column=14).value = ''
            current_sheet.cell(row=row, column=21).value = ''
        else:
            newBuildValue = current_sheet.cell(row=row, column=14).value
            print('Consist:',newBuildValue)
            #Searches the inbound columns to find where power came from
            foundPower = searchPower(newBuildValue, row)
            current_sheet.cell(row=row, column=21).value = foundPower
            print(current_sheet.cell(row=row, column=12).value,' changed to: ',newBuild, ' FROM:', current_sheet.cell(row=row, column=21).value, '.\n')
    else:
        print('No changes made.\n')
    menu()

def dispatchReport():
    print('\n---------- Inbound Trains ----------\n')
    inbound_train,inbound_power = (readMultColumnsTable(4, 15, 1, 3))
    table_format(inbound_train, inbound_power, 'Train Symbol','Power')
    print('\n---------- Outbound Trains ----------\n')
    outbound_trains, outbound_power = (readMultColumnsTable(4, 16, 12, 14))
    table_format(outbound_trains, outbound_power, 'Train Symbol','Power')
    print('\n---------- Extra Inbounds ----------\n')
    extra_inbounds, extra_inbound_power = (readMultColumnsTable(15, 28, 1, 3))
    table_format(extra_inbounds, extra_inbound_power, 'Train Symbol','Power')
    print('\n---------- Extra Outbounds ----------\n')
    extra_outbound, extra_outbound_power = (readMultColumnsTable(16, 28, 12, 14))
    table_format(extra_outbound, extra_outbound_power, 'Train Symbol','Power')
    print('\n')
    menu()
    
def fromBuilt():

    outbound_trains, outbound_power = (readMultColumnsTable(4, 18, 12, 14))
    from_symbol = readSingleColumn(4, 18, 21)
    #table_format(outbound_trains, outbound_power, 'Train Symbol', 'Power')
    table_format_three(outbound_trains, outbound_power, from_symbol, 'Train Symbol', 'Power', 'From')
    print('\n')
    #for i in range(4, 15):
    #    x = current_sheet.cell(row=i, column=12).value
    #    y = x[:3]
    #    z = current_sheet.cell(row=i, column=14).value
    #    a = current_sheet.cell(row=i, column=21).value
    #    print(y,': ',z)
    #    print('From:',a,'\n')
    #    #print('----: ',a,'\n')
    #print('\n')
    menu()

def savePowersheet():
    print('Saving workbook.....')
    #wb.save('/home/'+getpass.getuser()+'/Documents/test.xlsx')
    wb.template = False
    now = datetime. now()
    current_time = now. strftime("%H%M")
    wb.save('Enola Powersheet '+curDate+' '+current_time+'.xlsx')
    #wb.save('test.xlsx')

def newest(path):
    files = os.listdir(path)
    paths = [os.path.join(path, basename) for basename in files]
    return max(paths, key=os.path.getctime)
    #if basename.endswith('.csv')

def create_database():
    c.execute('CREATE TABLE IF NOT EXISTS enoladb (unit_number INTEGER UNIQUE,\
              model TEXT, cab_signals TEXT, lsl TEXT, fra_Date TEXT, epa_date\
              TEXT, lube_due TEXT, cs_due TEXT, afm_due TEXT, fuel_capacity\
              TEXT,incoming_fuel TEXT, current_fuel TEXT, ptc_status TEXT, \
              direction TEXT, updated TEXT, home_shop TEXT, alt_shop TEXT)')

def rundown():
    while True:
        unit = input('Input unit number(with or without direction): ')
        if unit == '' or unit == 'q' or unit == 'Q':
            print('Exiting')
            break
        else:
            try:
                unit_number = int(unit)
                direction = '-'
            except ValueError:
                unit_number = int(unit[:-1])
                direction = unit[-1]
            fuel_level = input('Fuel level: ')
            if direction == '-':
               fuel_db_update(unit_number, fuel_level)
            else:
                full_rundown(unit_number, fuel_level, direction)

def fuel_db_update(unit, fuel):
    c.execute("INSERT INTO enoladb (unit_number,\
        incoming_fuel, updated) VALUES (?,?,?)\
        ON CONFLICT(unit_number) DO UPDATE SET \
        incoming_fuel=excluded.incoming_fuel,\
        updated=excluded.updated;",\
        (unit,fuel,recorded_time))
    conn.commit()
 

def full_rundown(unit, fuel, direct):
    c.execute("INSERT INTO enoladb (unit_number,\
        incoming_fuel, direction, updated) VALUES (?,?,?,?)\
        ON CONFLICT(unit_number) DO UPDATE SET \
        incoming_fuel=excluded.incoming_fuel,\
        direction=excluded.direction,\
        updated=excluded.updated;",\
        (unit,fuel,direct,recorded_time))
    conn.commit()
           
if __name__ == '__main__':
    create_database()
    menu()
    conn.close()
